"""
Analytics and reporting for the grading system.

Generates class analytics, reports, and export data.
"""

import json
import csv
from typing import Dict, List, Optional
from pathlib import Path
from datetime import datetime
import numpy as np

from core.models import (
    GradingSession, GradedCopy, AnalyticsReport,
    ExportOptions
)


class AnalyticsGenerator:
    """
    Generates analytics reports for graded sessions.
    """

    def __init__(self, session: GradingSession):
        """
        Initialize analytics generator.

        Args:
            session: Grading session to analyze
        """
        self.session = session

    def generate(self) -> AnalyticsReport:
        """
        Generate complete analytics report.

        Returns:
            AnalyticsReport object
        """
        if not self.session.graded_copies:
            # Return report with default values when no copies were graded
            return AnalyticsReport(
                session_id=self.session.session_id,
                mean_score=0.0,
                median_score=0.0,
                std_dev=0.0,
                min_score=0.0,
                max_score=0.0
            )

        scores = [g.total_score for g in self.session.graded_copies]

        return AnalyticsReport(
            session_id=self.session.session_id,
            mean_score=float(np.mean(scores)),
            median_score=float(np.median(scores)),
            std_dev=float(np.std(scores)),
            min_score=float(min(scores)),
            max_score=float(max(scores)),
            score_distribution=self._compute_distribution(scores),
            question_stats=self._compute_question_stats(),
            common_errors=self._get_common_errors(),
            exceptional_answers=self._get_exceptional_answers(),
            generated_at=datetime.now()
        )

    def _compute_distribution(self, scores: List[float]) -> Dict[str, int]:
        """Compute score distribution buckets."""
        distribution = {}

        # Define buckets based on max score
        if self.session.graded_copies:
            max_score = self.session.graded_copies[0].max_score
            bucket_size = max_score / 5
        else:
            max_score = 20
            bucket_size = 4

        for i in range(5):
            lower = i * bucket_size
            upper = (i + 1) * bucket_size
            key = f"{lower:.0f}-{upper:.0f}"

            count = sum(1 for s in scores if lower <= s < upper)
            distribution[key] = count

        return distribution

    def _compute_question_stats(self) -> Dict[str, Dict[str, float]]:
        """Compute statistics per question."""
        stats = {}

        # Get all question IDs
        question_ids = set()
        for graded in self.session.graded_copies:
            question_ids.update(graded.grades.keys())

        for q_id in question_ids:
            scores = [
                g.grades.get(q_id, 0)
                for g in self.session.graded_copies
                if q_id in g.grades
            ]

            if scores:
                stats[q_id] = {
                    "mean": float(np.mean(scores)),
                    "median": float(np.median(scores)),
                    "min": float(min(scores)),
                    "max": float(max(scores)),
                    "difficulty": 1.0 - (np.mean(scores) / 5.0) if scores else 0.5
                }

        return stats

    def _get_common_errors(self) -> List[str]:
        """Get common errors from class map."""
        errors = []

        if self.session.class_map:
            for q_id, analysis in self.session.class_map.question_analyses.items():
                for error in analysis.common_errors[:2]:
                    errors.append(f"Q{q_id}: {error}")

        return errors

    def _get_exceptional_answers(self) -> List[str]:
        """Get exceptional/high-scoring answers."""
        exceptional = []

        for graded in self.session.graded_copies:
            if graded.total_score >= graded.max_score * 0.95:
                copy = next(
                    (c for c in self.session.copies if c.id == graded.copy_id),
                    None
                )
                if copy:
                    name = copy.student_name or "Anonymous"
                    exceptional.append(
                        f"{name}: {graded.total_score}/{graded.max_score}"
                    )

        return exceptional


class DataExporter:
    """
    Exports session data in various formats.
    """

    def __init__(self, session: GradingSession, output_dir: str = None):
        """
        Initialize data exporter.

        Args:
            session: Grading session
            output_dir: Output directory (default: outputs/reports)
        """
        self.session = session
        self.output_dir = Path(output_dir or "outputs/reports")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_json(self, filename: str = None) -> str:
        """
        Export session data as comprehensive JSON.

        Args:
            filename: Output filename (auto-generated if None)

        Returns:
            Path to exported file
        """
        if filename is None:
            filename = f"{self.session.session_id}_export.json"

        output_path = self.output_dir / filename

        # Build comprehensive export
        data = {
            "session": {
                "id": self.session.session_id,
                "created_at": self.session.created_at.isoformat(),
                "status": self.session.status,
                "total_copies": len(self.session.copies),
                "graded_copies": len(self.session.graded_copies)
            },
            "bareme": self._build_bareme(),
            "students": self._build_students_data(),
            "summary": self._build_summary()
        }

        # Add grading audit summary if available from any graded copy
        for graded in self.session.graded_copies:
            if graded.grading_audit:
                data["grading_audit_summary"] = {
                    "mode": graded.grading_audit.mode,
                    "grading_method": graded.grading_audit.grading_method,
                    "verification_mode": graded.grading_audit.verification_mode,
                    "providers": [
                        {"id": p.id, "model": p.model}
                        for p in graded.grading_audit.providers
                    ],
                    "summary": {
                        "total_questions": graded.grading_audit.summary.total_questions,
                        "agreed_initial": graded.grading_audit.summary.agreed_initial,
                        "required_verification": graded.grading_audit.summary.required_verification,
                        "required_ultimatum": graded.grading_audit.summary.required_ultimatum,
                        "final_agreement_rate": graded.grading_audit.summary.final_agreement_rate
                    }
                }
                break

        # Write JSON with proper formatting
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str, ensure_ascii=False)

        return str(output_path)

    def export_individual_copies(self) -> List[str]:
        """
        Export each student's graded copy as individual JSON file.

        Returns:
            List of paths to exported files
        """
        exported_paths = []

        for graded in self.session.graded_copies:
            # Find corresponding copy
            copy = next(
                (c for c in self.session.copies if c.id == graded.copy_id),
                None
            )

            # Build student name for filename (sanitize)
            student_name = copy.student_name if copy and copy.student_name else "Anonyme"
            safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in student_name)
            safe_name = safe_name.replace(' ', '_')[:50]  # Limit length

            filename = f"{safe_name}_{graded.copy_id}.json"
            output_path = self.output_dir / filename

            # Build individual export data
            data = {
                "session_id": self.session.session_id,
                "copy_id": graded.copy_id,
                "student_name": student_name,
                "graded_at": graded.graded_at.isoformat() if hasattr(graded, 'graded_at') else None,
                "total_score": graded.total_score,
                "max_score": graded.max_score,
                "percentage": round(graded.total_score / graded.max_score * 100, 1) if graded.max_score > 0 else 0,
                "overall_confidence": graded.confidence,
                "questions": {},
                "feedback": graded.feedback or ""
            }

            # Add per-question details
            for q_id, score in graded.grades.items():
                q_confidence = graded.confidence_by_question.get(q_id, graded.confidence)

                # Get student feedback
                feedback = ""
                if graded.student_feedback:
                    if isinstance(graded.student_feedback, dict):
                        feedback = graded.student_feedback.get(q_id, "")

                detected_answer = ""
                if copy and copy.content_summary:
                    detected_answer = copy.content_summary.get(q_id, "")

                data["questions"][q_id] = {
                    "score": score,
                    "max_points": self._get_question_max_points(q_id),
                    "confidence": q_confidence,
                    "detected_answer": detected_answer,
                    "feedback": feedback
                }

            # Write JSON
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, default=str, ensure_ascii=False)

            exported_paths.append(str(output_path))

        return exported_paths

    def _build_bareme(self) -> Dict:
        """Build barème (grading scale) section."""
        bareme = {}

        # Get scale from policy.question_weights (the correct source)
        if hasattr(self.session, 'policy') and self.session.policy and self.session.policy.question_weights:
            for q_id, max_pts in self.session.policy.question_weights.items():
                bareme[q_id] = {
                    "max_points": max_pts,
                    "description": ""
                }

        return bareme

    def _build_students_data(self) -> List[Dict]:
        """Build detailed student data."""
        students = []

        for graded in self.session.graded_copies:
            # Find corresponding copy
            copy = next(
                (c for c in self.session.copies if c.id == graded.copy_id),
                None
            )

            student = {
                "copy_id": graded.copy_id,
                "name": copy.student_name if copy else "Anonyme",
                "total_score": graded.total_score,
                "max_score": graded.max_score,
                "percentage": round(graded.total_score / graded.max_score * 100, 1) if graded.max_score > 0 else 0,
                "overall_confidence": graded.confidence,
                "questions": {},
                "feedback": graded.feedback or ""
            }

            # Add per-question details
            for q_id, score in graded.grades.items():
                # Get question-specific confidence
                q_confidence = graded.confidence_by_question.get(q_id, graded.confidence)

                # Get student feedback
                feedback = ""
                if graded.student_feedback:
                    if isinstance(graded.student_feedback, dict):
                        feedback = graded.student_feedback.get(q_id, "")

                # Get detected answer from copy
                detected_answer = ""
                if copy and copy.content_summary:
                    detected_answer = copy.content_summary.get(q_id, "")

                student["questions"][q_id] = {
                    "score": score,
                    "max_points": self._get_question_max_points(q_id),
                    "confidence": q_confidence,
                    "detected_answer": detected_answer,
                    "feedback": feedback
                }

            # Add grading audit trail if available
            if graded.grading_audit:
                student["grading_audit"] = {
                    "mode": graded.grading_audit.mode,
                    "grading_method": graded.grading_audit.grading_method,
                    "verification_mode": graded.grading_audit.verification_mode,
                    "providers": [
                        {"id": p.id, "model": p.model}
                        for p in graded.grading_audit.providers
                    ],
                    "summary": {
                        "total_questions": graded.grading_audit.summary.total_questions,
                        "agreed_initial": graded.grading_audit.summary.agreed_initial,
                        "final_agreement_rate": graded.grading_audit.summary.final_agreement_rate
                    }
                }

            students.append(student)

        return students

    def _get_question_max_points(self, question_id: str) -> float:
        """Get max points for a question from the session policy."""
        # Use policy.question_weights (the authoritative source)
        if hasattr(self.session, 'policy') and self.session.policy and self.session.policy.question_weights:
            return self.session.policy.question_weights.get(question_id, 0)

        return 0

    def _build_summary(self) -> Dict:
        """Build summary statistics."""
        if not self.session.graded_copies:
            return {}

        scores = [g.total_score for g in self.session.graded_copies]
        max_score = self.session.graded_copies[0].max_score if self.session.graded_copies else 1

        return {
            "average_score": round(float(np.mean(scores)), 2),
            "median_score": round(float(np.median(scores)), 2),
            "min_score": round(float(min(scores)), 2),
            "max_score": round(float(max(scores)), 2),
            "std_deviation": round(float(np.std(scores)), 2),
            "average_percentage": round(float(np.mean(scores)) / max_score * 100, 1) if max_score > 0 else 0,
            "question_stats": self._build_question_stats_summary()
        }

    def _build_question_stats_summary(self) -> Dict:
        """Build per-question statistics."""
        stats = {}

        # Get all question IDs
        question_ids = set()
        for graded in self.session.graded_copies:
            question_ids.update(graded.grades.keys())

        for q_id in sorted(question_ids):
            scores = []
            confidences = []

            for graded in self.session.graded_copies:
                if q_id in graded.grades:
                    scores.append(graded.grades[q_id])
                    if q_id in graded.confidence_by_question:
                        confidences.append(graded.confidence_by_question[q_id])

            if scores:
                max_pts = self._get_question_max_points(q_id)
                success_rate = round(float(np.mean(scores)) / max_pts * 100, 1) if max_pts > 0 else 0
                stats[q_id] = {
                    "average": round(float(np.mean(scores)), 2),
                    "max_points": max_pts,
                    "success_rate": min(success_rate, 100.0),  # Cap at 100%
                    "average_confidence": round(float(np.mean(confidences)), 2) if confidences else 0.5
                }

        return stats

    def export_csv(self, filename: str = None) -> str:
        """
        Export grades as CSV.

        Args:
            filename: Output filename (auto-generated if None)

        Returns:
            Path to exported file
        """
        if filename is None:
            filename = f"{self.session.session_id}_grades.csv"

        output_path = self.output_dir / filename

        # Get all question IDs
        question_ids = set()
        for graded in self.session.graded_copies:
            question_ids.update(graded.grades.keys())
        question_ids = sorted(question_ids)

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            header = ["élève", "total", "max"]
            header.extend([f"q_{qid}" for qid in question_ids])
            header.append("commentaire")
            writer.writerow(header)

            # Rows
            for graded in self.session.graded_copies:
                copy = next(
                    (c for c in self.session.copies if c.id == graded.copy_id),
                    None
                )

                row = [
                    copy.student_name if copy else "",
                    graded.total_score,
                    graded.max_score
                ]

                for qid in question_ids:
                    row.append(graded.grades.get(qid, ""))

                row.append(graded.feedback or "")
                writer.writerow(row)

        return str(output_path)

    def export_excel(self, filename: str = None) -> str:
        """
        Export session results as formatted Excel file.

        Creates a professional Excel file with:
        - Styled header row (bold, colored background)
        - Auto-adjusted column widths
        - All student data: names, per-question grades, totals, feedback

        Args:
            filename: Output filename (auto-generated if None)

        Returns:
            Path to exported Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        if filename is None:
            date_str = datetime.now().strftime("%Y-%m-%d")
            filename = f"correction_{date_str}_{self.session.session_id[:8]}.xlsx"

        output_path = self.output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"

        # Define header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Get all question IDs (sorted)
        question_ids = set()
        for graded in self.session.graded_copies:
            question_ids.update(graded.grades.keys())
        question_ids = sorted(question_ids, key=lambda x: (len(x), x))

        # Write header row
        headers = ["Nom élève", "Total", "Note maximale"]
        headers.extend([f"Q{qid}" for qid in question_ids])
        headers.append("Appréciation")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_num, graded in enumerate(self.session.graded_copies, 2):
            # Find student name from copy
            copy = next((c for c in self.session.copies if c.id == graded.copy_id), None)
            student_name = copy.student_name if copy else "Anonyme"

            # Basic info
            ws.cell(row=row_num, column=1, value=student_name)
            ws.cell(row=row_num, column=2, value=graded.total_score)
            ws.cell(row=row_num, column=3, value=graded.max_score)

            # Per-question grades
            for col_num, qid in enumerate(question_ids, 4):
                grade = graded.grades.get(qid, "")
                ws.cell(row=row_num, column=col_num, value=grade)

            # Feedback
            feedback_col = 4 + len(question_ids)
            ws.cell(row=row_num, column=feedback_col, value=graded.feedback or "")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save file
        wb.save(output_path)

        return str(output_path)

    def export_analytics(self, filename: str = None) -> str:
        """
        Export analytics report as JSON.

        Args:
            filename: Output filename (auto-generated if None)

        Returns:
            Path to exported file
        """
        if filename is None:
            filename = f"{self.session.session_id}_analytics.json"

        output_path = self.output_dir / filename

        generator = AnalyticsGenerator(self.session)
        report = generator.generate()

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report.model_dump(mode='json'), f, indent=2, default=str)

        return str(output_path)

    def export_all(self, options: ExportOptions = None) -> Dict[str, str]:
        """
        Export all data in specified formats.

        Args:
            options: ExportOptions

        Returns:
            Dict of format -> file path
        """
        if options is None:
            options = ExportOptions()

        results = {}

        if options.format in ["json", "all"]:
            results["json"] = self.export_json()

        if options.format in ["csv", "all"]:
            results["csv"] = self.export_csv()

        if options.format in ["json", "all"] and options.include_analytics:
            results["analytics"] = self.export_analytics()

        return results


def export_session(
    session: GradingSession,
    formats: str = "all",
    output_dir: str = None
) -> Dict[str, str]:
    """
    Convenience function to export a session.

    Args:
        session: Grading session to export
        formats: Comma-separated formats (json, csv, analytics)
        output_dir: Output directory

    Returns:
        Dict of format -> file path
    """
    exporter = DataExporter(session, output_dir)
    results = {}

    for fmt in formats.split(","):
        fmt = fmt.strip()
        if fmt == "json":
            results["json"] = exporter.export_json()
        elif fmt == "csv":
            results["csv"] = exporter.export_csv()
        elif fmt == "analytics":
            results["analytics"] = exporter.export_analytics()

    return results
